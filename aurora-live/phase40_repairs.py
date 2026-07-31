from __future__ import annotations

import io
import os
import re
import urllib.parse
from datetime import datetime, timezone
from typing import Any

from phase40_markets import (
    AlphaVantageEquitiesAdapter,
    BaseAdapter,
    CoinbaseCryptoAdapter,
    ECBFXAdapter,
    EIAEnergyAdapter,
    KalshiPredictionMarketsAdapter,
    MarketCoordinator,
    MarketObservation,
    WorldBankCommoditiesAdapter,
    WorldBankIndicatorsAdapter,
    _event_time,
    _float,
    _json,
    _now,
    _parse_time,
    _request,
)


def _normalized(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", str(value or "").strip().upper()).strip("_")


def _date_like(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        current = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
        return current.astimezone(timezone.utc)
    parsed = _parse_time(value)
    if parsed is not None:
        return parsed
    text = str(value or "").strip()
    match = re.fullmatch(r"(19|20)\d{2}[Mm-](0[1-9]|1[0-2])", text)
    if match:
        year = int(text[:4])
        month = int(text[-2:])
        return datetime(year, month, 1, tzinfo=timezone.utc)
    return None


class ResilientWorldBankCommoditiesAdapter(WorldBankCommoditiesAdapter):
    """Parse current and legacy Pink Sheet workbooks with offset, multi-row headers."""

    aliases = {
        "CRUDE_BRENT": {"CRUDE_BRENT", "OIL_BRENT", "BRENT"},
        "GOLD": {"GOLD"},
        "SILVER": {"SILVER"},
        "COPPER": {"COPPER"},
        "NGAS_US": {"NGAS_US", "NATURAL_GAS_US", "US_NATURAL_GAS"},
        "COAL_AUS": {"COAL_AUS", "AUSTRALIAN_COAL", "COAL_AUSTRALIA"},
    }

    def selected(self) -> list[str]:
        raw = str(os.getenv(self.series_env) or "CRUDE_BRENT,GOLD,SILVER,COPPER,NGAS_US,COAL_AUS")
        return [_normalized(item) for item in raw.split(",") if _normalized(item)][:50]

    @classmethod
    def _matches(cls, selected: set[str], labels: list[str]) -> bool:
        if not selected:
            return True
        normalized_labels = {_normalized(label) for label in labels if _normalized(label)}
        joined = "_".join(sorted(normalized_labels))
        for requested in selected:
            candidates = cls.aliases.get(requested, {requested})
            for candidate in candidates:
                if candidate in normalized_labels or candidate in joined:
                    return True
        return False

    @staticmethod
    def _instrument(labels: list[str], index: int) -> str:
        normalized = [_normalized(label) for label in labels if _normalized(label)]
        code_like = [label for label in normalized if re.fullmatch(r"[A-Z][A-Z0-9_]{2,}", label)]
        return (code_like[-1] if code_like else (normalized[0] if normalized else f"SERIES_{index}"))[:160]

    def fetch(self, timeout: int = 45) -> list[MarketObservation]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for World Bank commodity ingestion") from exc
        try:
            content = _request(
                self.source_url(),
                timeout=timeout,
                accept="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise RuntimeError("World Bank Pink Sheet request or parse failed") from exc

        sheet = next((workbook[name] for name in workbook.sheetnames if "monthly" in name.lower()), workbook.active)
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise RuntimeError("World Bank Pink Sheet returned an empty workbook")

        header_row = None
        date_column = None
        for row_index, row in enumerate(rows[:40]):
            for column_index, value in enumerate(row):
                token = _normalized(value)
                if token in {"DATE", "MONTH", "PERIOD"} or token.endswith("_DATE"):
                    header_row = row_index
                    date_column = column_index
                    break
            if header_row is not None:
                break

        if header_row is None or date_column is None:
            for row_index, row in enumerate(rows[:80]):
                for column_index, value in enumerate(row[:12]):
                    if _date_like(value) is not None and sum(_float(cell) is not None for cell in row) >= 2:
                        header_row = max(0, row_index - 3)
                        date_column = column_index
                        break
                if date_column is not None:
                    break
        if header_row is None or date_column is None:
            raise RuntimeError("World Bank Pink Sheet date column was not found")

        data_rows: list[tuple[datetime, tuple[Any, ...]]] = []
        for row in rows[header_row + 1 :]:
            if date_column >= len(row):
                continue
            parsed = _date_like(row[date_column])
            if parsed is None:
                continue
            if sum(_float(cell) is not None for cell in row) < 2:
                continue
            data_rows.append((parsed, row))
        if not data_rows:
            raise RuntimeError("World Bank Pink Sheet returned no dated numeric rows")
        latest_date, latest = max(data_rows, key=lambda item: item[0])

        header_start = max(0, header_row - 2)
        header_end = min(len(rows), header_row + 5)
        selected = set(self.selected())
        observed_at = _now()
        event = _event_time(latest_date, observed_at)
        output: list[MarketObservation] = []
        max_columns = max(len(row) for row in rows[header_start:header_end])
        for column_index in range(max_columns):
            if column_index == date_column or column_index >= len(latest):
                continue
            value = _float(latest[column_index])
            if value is None:
                continue
            labels = [
                str(rows[row_index][column_index]).strip()
                for row_index in range(header_start, header_end)
                if column_index < len(rows[row_index]) and rows[row_index][column_index] not in (None, "")
            ]
            if not self._matches(selected, labels):
                continue
            instrument = self._instrument(labels, column_index)
            unit = next(
                (
                    label
                    for label in labels
                    if any(marker in label.lower() for marker in ("$", "usd", "ton", "bbl", "kg", "mmbtu", "toz", "cents"))
                ),
                "published unit",
            )
            output.append(
                MarketObservation(
                    self.domain,
                    self.name,
                    instrument,
                    f"{instrument}:{event}",
                    observed_at,
                    event,
                    value,
                    unit,
                    "official-monthly-price",
                    {"series": instrument, "period": event, "value": value, "unit": unit, "header_labels": labels},
                    {
                        "provider": "World Bank Prospects Group",
                        "source": self.source_url(),
                        "scope_env": self.series_env,
                        "workbook_sheet": sheet.title,
                        "date_column_index": date_column,
                        "license_note": self.license_note,
                    },
                )
            )
        if not output:
            raise RuntimeError("World Bank Pink Sheet contained no selected numeric series")
        return output


class ResilientWorldBankIndicatorsAdapter(WorldBankIndicatorsAdapter):
    """Query countries independently so one unavailable combination cannot poison the domain."""

    def fetch(self, timeout: int = 30) -> list[MarketObservation]:
        observed_at = _now()
        output: list[MarketObservation] = []
        failures: list[str] = []
        for indicator in self.indicators():
            for country in self.countries():
                url = (
                    f"{self.endpoint}/country/{urllib.parse.quote(country, safe='')}/indicator/"
                    f"{urllib.parse.quote(indicator, safe='.') }?"
                    f"{urllib.parse.urlencode({'format': 'json', 'per_page': 10, 'mrnev': 1})}"
                )
                try:
                    payload = _json(url, timeout=timeout)
                except Exception:
                    failures.append(f"{country}:{indicator}")
                    continue
                rows = payload[1] if isinstance(payload, list) and len(payload) > 1 and isinstance(payload[1], list) else []
                for row in rows:
                    if not isinstance(row, dict):
                        continue
                    value = _float(row.get("value"))
                    if value is None:
                        continue
                    country_code = str(row.get("countryiso3code") or (row.get("country") or {}).get("id") or country).upper()
                    code = str((row.get("indicator") or {}).get("id") or indicator)
                    period = str(row.get("date") or "")
                    if not country_code or not period:
                        continue
                    event = _event_time(period, observed_at)
                    instrument = f"{country_code}:{code}"
                    output.append(
                        MarketObservation(
                            self.domain,
                            self.name,
                            instrument,
                            f"{instrument}:{event}",
                            observed_at,
                            event,
                            value,
                            "indicator-defined unit",
                            "official-statistic",
                            row,
                            {
                                "provider": "World Bank",
                                "source": self.endpoint,
                                "country": country_code,
                                "indicator": code,
                                "countries_env": self.countries_env,
                                "indicators_env": self.indicators_env,
                                "request_url_persisted": False,
                                "license_note": self.license_note,
                            },
                        )
                    )
                    break
        if not output:
            detail = f" ({len(failures)} scoped requests failed)" if failures else ""
            raise RuntimeError(f"World Bank indicators returned no numeric observations{detail}")
        return output


class ProductionMarketCoordinator(MarketCoordinator):
    def __init__(self, store) -> None:
        self.store = store
        self.adapters: list[BaseAdapter] = [
            AlphaVantageEquitiesAdapter(),
            EIAEnergyAdapter(),
            ResilientWorldBankCommoditiesAdapter(),
            ECBFXAdapter(),
            CoinbaseCryptoAdapter(),
            ResilientWorldBankIndicatorsAdapter(),
            KalshiPredictionMarketsAdapter(),
        ]
        self._ensure_registered()

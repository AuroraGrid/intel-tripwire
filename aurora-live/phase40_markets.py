from __future__ import annotations

import csv
import io
import json
import os
import sqlite3
import threading
import time
import urllib.parse
import urllib.request
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

DOMAINS = (
    "equities",
    "energy",
    "commodities",
    "fx",
    "crypto",
    "economic_indicators",
    "prediction_markets",
)
PROVIDER_STATES = {"ONLINE", "DEGRADED", "OFFLINE", "NOT_CONFIGURED"}
USER_AGENT = "AURORA-LIVE/1.0 (+https://github.com/hr185882-creator/intel-tripwire)"


def _now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _parse_time(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    if len(text) == 4 and text.isdigit():
        text = f"{text}-01-01T00:00:00Z"
    elif len(text) == 7 and text[4] in {"-", "M"}:
        text = f"{text.replace('M', '-')}-01T00:00:00Z"
    elif len(text) == 10 and text[4] == "-":
        text = f"{text}T00:00:00Z"
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).astimezone(timezone.utc)
    except ValueError:
        return None


def _event_time(value: Any, fallback: str) -> str:
    if isinstance(value, datetime):
        current = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
        return current.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")
    parsed = _parse_time(value)
    return parsed.isoformat().replace("+00:00", "Z") if parsed else fallback


def _float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _is_postgres(target: str) -> bool:
    return target.startswith(("postgresql://", "postgres://"))


def _request(url: str, *, timeout: int = 30, accept: str = "application/json") -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, "Accept": accept})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def _json(url: str, *, timeout: int = 30) -> Any:
    return json.loads(_request(url, timeout=timeout).decode("utf-8", "replace"))


@dataclass(frozen=True)
class MarketObservation:
    domain: str
    provider: str
    instrument: str
    external_id: str
    observed_at: str
    event_time: str
    value: float
    unit: str
    status: str
    payload: dict[str, Any]
    provenance: dict[str, Any]

    def value_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class ProviderRun:
    provider: str
    domain: str
    configured: bool
    successful: bool
    observations: int
    error: str = ""
    started_at: str = ""
    completed_at: str = ""
    duration_ms: int = 0

    def value(self) -> dict[str, Any]:
        return asdict(self)


class MarketStore:
    """Durable SQLite/PostgreSQL market observations, provider health and run history."""

    def __init__(self, target: str = ":memory:") -> None:
        self.target = str(target)
        self.postgres = _is_postgres(self.target)
        self._lock = threading.RLock()
        if self.postgres:
            try:
                import psycopg
                from psycopg.rows import dict_row
            except ImportError as exc:
                raise RuntimeError("psycopg is required for PostgreSQL market storage") from exc
            self._connection = psycopg.connect(self.target, row_factory=dict_row)
            self._p = "%s"
        else:
            if self.target != ":memory:":
                Path(self.target).parent.mkdir(parents=True, exist_ok=True)
            self._connection = sqlite3.connect(self.target, check_same_thread=False)
            self._connection.row_factory = sqlite3.Row
            self._p = "?"
        self._initialize()

    def _initialize(self) -> None:
        observation_pk = "observation_id BIGSERIAL PRIMARY KEY" if self.postgres else "observation_id INTEGER PRIMARY KEY AUTOINCREMENT"
        run_pk = "run_id BIGSERIAL PRIMARY KEY" if self.postgres else "run_id INTEGER PRIMARY KEY AUTOINCREMENT"
        statements = [
            f"""CREATE TABLE IF NOT EXISTS market_observations (
                {observation_pk}, domain TEXT NOT NULL, provider TEXT NOT NULL,
                instrument TEXT NOT NULL, external_id TEXT NOT NULL,
                observed_at TEXT NOT NULL, event_time TEXT NOT NULL,
                value DOUBLE PRECISION NOT NULL, unit TEXT NOT NULL,
                status TEXT NOT NULL, payload_json TEXT NOT NULL,
                provenance_json TEXT NOT NULL)""",
            """CREATE TABLE IF NOT EXISTS market_provider_health (
                provider TEXT PRIMARY KEY, domain TEXT NOT NULL, state TEXT NOT NULL,
                last_attempt_at TEXT NOT NULL, last_success_at TEXT NOT NULL,
                consecutive_failures INTEGER NOT NULL, event_age_seconds INTEGER NOT NULL,
                last_error TEXT NOT NULL, completeness_note TEXT NOT NULL,
                license_note TEXT NOT NULL, updated_at TEXT NOT NULL)""",
            f"""CREATE TABLE IF NOT EXISTS market_provider_runs (
                {run_pk}, provider TEXT NOT NULL, domain TEXT NOT NULL,
                configured INTEGER NOT NULL, successful INTEGER NOT NULL,
                observations INTEGER NOT NULL, started_at TEXT NOT NULL,
                completed_at TEXT NOT NULL, duration_ms INTEGER NOT NULL,
                error TEXT NOT NULL, metadata_json TEXT NOT NULL)""",
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_market_observation_identity ON market_observations(domain,provider,external_id)",
            "CREATE INDEX IF NOT EXISTS idx_market_domain_time ON market_observations(domain,observation_id DESC)",
            "CREATE INDEX IF NOT EXISTS idx_market_provider_time ON market_observations(provider,observation_id DESC)",
            "CREATE INDEX IF NOT EXISTS idx_market_runs_provider ON market_provider_runs(provider,run_id DESC)",
        ]
        with self._lock:
            cursor = self._connection.cursor()
            if self.postgres:
                cursor.execute("SELECT pg_advisory_xact_lock(%s)", (400040,))
            for statement in statements:
                cursor.execute(statement)
            self._connection.commit()

    @staticmethod
    def _dict(row: Any) -> dict[str, Any]:
        return dict(row) if row is not None else {}

    @staticmethod
    def _id(row: Any, key: str) -> int:
        if isinstance(row, dict):
            return int(row[key])
        try:
            return int(row[key])
        except (TypeError, KeyError, IndexError):
            return int(row[0])

    def record(self, observation: MarketObservation) -> int:
        if observation.domain not in DOMAINS:
            raise ValueError("invalid market domain")
        if not observation.provider or not observation.instrument or not observation.external_id:
            raise ValueError("provider, instrument and external_id are required")
        values = (
            observation.domain,
            observation.provider,
            observation.instrument,
            observation.external_id,
            observation.observed_at,
            observation.event_time,
            float(observation.value),
            observation.unit,
            observation.status,
            json.dumps(observation.payload, sort_keys=True, separators=(",", ":"), default=str),
            json.dumps(observation.provenance, sort_keys=True, separators=(",", ":"), default=str),
        )
        with self._lock:
            cursor = self._connection.cursor()
            cursor.execute(
                f"SELECT observation_id FROM market_observations WHERE domain={self._p} AND provider={self._p} AND external_id={self._p} LIMIT 1",
                (observation.domain, observation.provider, observation.external_id),
            )
            existing = cursor.fetchone()
            if existing is not None:
                return self._id(existing, "observation_id")
            if self.postgres:
                cursor.execute(
                    "INSERT INTO market_observations(domain,provider,instrument,external_id,observed_at,event_time,value,unit,status,payload_json,provenance_json) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING observation_id",
                    values,
                )
                identifier = self._id(cursor.fetchone(), "observation_id")
            else:
                cursor.execute(
                    "INSERT INTO market_observations(domain,provider,instrument,external_id,observed_at,event_time,value,unit,status,payload_json,provenance_json) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    values,
                )
                identifier = int(cursor.lastrowid)
            self._connection.commit()
            return identifier

    def observations(self, domain: str = "", provider: str = "", instrument: str = "", limit: int = 250) -> list[dict[str, Any]]:
        if domain and domain not in DOMAINS:
            raise ValueError("invalid market domain")
        clauses: list[str] = []
        values: list[Any] = []
        if domain:
            clauses.append(f"domain={self._p}")
            values.append(domain)
        if provider:
            clauses.append(f"provider={self._p}")
            values.append(provider)
        if instrument:
            clauses.append(f"instrument={self._p}")
            values.append(instrument)
        where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
        values.append(max(1, min(int(limit), 5000)))
        with self._lock:
            rows = self._connection.execute(
                f"SELECT * FROM market_observations{where} ORDER BY observation_id DESC LIMIT {self._p}",
                tuple(values),
            ).fetchall()
        output: list[dict[str, Any]] = []
        for row in rows:
            item = self._dict(row)
            for source, target in (("payload_json", "payload"), ("provenance_json", "provenance")):
                try:
                    item[target] = json.loads(item.pop(source, "{}"))
                except json.JSONDecodeError:
                    item[target] = {}
            output.append(item)
        return output

    def upsert_provider(self, value: dict[str, Any]) -> None:
        domain = str(value.get("domain") or "")
        state = str(value.get("state") or "")
        if domain not in DOMAINS:
            raise ValueError("invalid market domain")
        if state not in PROVIDER_STATES:
            raise ValueError("invalid provider state")
        columns = (
            "provider", "domain", "state", "last_attempt_at", "last_success_at",
            "consecutive_failures", "event_age_seconds", "last_error",
            "completeness_note", "license_note", "updated_at",
        )
        payload = {**value, "updated_at": value.get("updated_at") or _now()}
        updates = ",".join(f"{column}=excluded.{column}" for column in columns[1:])
        sql = f"INSERT INTO market_provider_health({','.join(columns)}) VALUES ({','.join([self._p] * len(columns))}) ON CONFLICT(provider) DO UPDATE SET {updates}"
        with self._lock:
            self._connection.execute(sql, tuple(payload.get(column, "") for column in columns))
            self._connection.commit()

    def providers(self, domain: str = "") -> list[dict[str, Any]]:
        if domain and domain not in DOMAINS:
            raise ValueError("invalid market domain")
        with self._lock:
            if domain:
                rows = self._connection.execute(
                    f"SELECT * FROM market_provider_health WHERE domain={self._p} ORDER BY provider",
                    (domain,),
                ).fetchall()
            else:
                rows = self._connection.execute("SELECT * FROM market_provider_health ORDER BY domain,provider").fetchall()
        return [self._dict(row) for row in rows]

    def record_run(self, run: ProviderRun, metadata: dict[str, Any] | None = None) -> int:
        values = (
            run.provider,
            run.domain,
            1 if run.configured else 0,
            1 if run.successful else 0,
            max(0, int(run.observations)),
            run.started_at or _now(),
            run.completed_at or _now(),
            max(0, int(run.duration_ms)),
            run.error,
            json.dumps(metadata or {}, sort_keys=True, separators=(",", ":"), default=str),
        )
        with self._lock:
            cursor = self._connection.cursor()
            if self.postgres:
                cursor.execute(
                    "INSERT INTO market_provider_runs(provider,domain,configured,successful,observations,started_at,completed_at,duration_ms,error,metadata_json) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING run_id",
                    values,
                )
                identifier = self._id(cursor.fetchone(), "run_id")
            else:
                cursor.execute(
                    "INSERT INTO market_provider_runs(provider,domain,configured,successful,observations,started_at,completed_at,duration_ms,error,metadata_json) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    values,
                )
                identifier = int(cursor.lastrowid)
            self._connection.commit()
            return identifier

    def runs(self, domain: str = "", provider: str = "", limit: int = 100) -> list[dict[str, Any]]:
        if domain and domain not in DOMAINS:
            raise ValueError("invalid market domain")
        clauses: list[str] = []
        values: list[Any] = []
        if domain:
            clauses.append(f"domain={self._p}")
            values.append(domain)
        if provider:
            clauses.append(f"provider={self._p}")
            values.append(provider)
        where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
        values.append(max(1, min(int(limit), 1000)))
        with self._lock:
            rows = self._connection.execute(
                f"SELECT * FROM market_provider_runs{where} ORDER BY run_id DESC LIMIT {self._p}",
                tuple(values),
            ).fetchall()
        output: list[dict[str, Any]] = []
        for row in rows:
            item = self._dict(row)
            item["configured"] = bool(item.get("configured"))
            item["successful"] = bool(item.get("successful"))
            try:
                item["metadata"] = json.loads(item.pop("metadata_json", "{}"))
            except json.JSONDecodeError:
                item["metadata"] = {}
            output.append(item)
        return output

    def health(self, max_age_seconds: int | None = None) -> dict[str, Any]:
        configured_age = os.getenv("AURORA_MARKETS_STALE_SECONDS") or "3600"
        maximum_age = max(60, int(max_age_seconds or configured_age))
        now = datetime.now(timezone.utc)
        provider_health: list[dict[str, Any]] = []
        for provider in self.providers():
            last_success = _parse_time(provider.get("last_success_at"))
            retrieval_age = int((now - last_success).total_seconds()) if last_success else None
            observations = self.observations(provider=provider["provider"], limit=1)
            runs = self.runs(provider=provider["provider"], limit=1)
            latest_run = runs[0] if runs else None
            fresh = bool(provider["state"] == "ONLINE" and retrieval_age is not None and retrieval_age <= maximum_age)
            operational = bool(fresh and observations and latest_run and latest_run["successful"] and latest_run["observations"] > 0)
            provider_health.append(
                {
                    **provider,
                    "seconds_since_success": retrieval_age,
                    "retrieval_fresh": fresh,
                    "event_age_seconds": max(0, int(provider.get("event_age_seconds") or 0)),
                    "operational": operational,
                    "observations": 1 if observations else 0,
                    "latest_observation": observations[0] if observations else None,
                    "latest_run": latest_run,
                }
            )
        domains = []
        for domain in DOMAINS:
            rows = [row for row in provider_health if row["domain"] == domain]
            domains.append(
                {
                    "domain": domain,
                    "providers": len(rows),
                    "operational_providers": sum(1 for row in rows if row["operational"]),
                    "operational": any(row["operational"] for row in rows),
                }
            )
        return {
            "max_age_seconds": maximum_age,
            "freshness_basis": "recent successful provider retrieval with durable numeric observations; market or publication age is reported separately",
            "providers": provider_health,
            "domains": domains,
            "operational_domains": sum(1 for row in domains if row["operational"]),
            "fully_operational": all(row["operational"] for row in domains),
            "generated_at": _now(),
        }

    def coverage(self) -> dict[str, Any]:
        health = self.health()
        return {
            "requirement": "each market domain requires a configured provider, recent successful retrieval and durable numeric observations",
            "domains": health["domains"],
            "qualified_domains": health["operational_domains"],
            "fully_qualified": health["fully_operational"],
            "generated_at": health["generated_at"],
        }


class BaseAdapter:
    name = ""
    domain = ""
    endpoint = ""
    license_note = "Provider terms apply."
    completeness_note = "Coverage is provider-dependent and is not assumed comprehensive or real-time."

    def configured(self) -> bool:
        return True

    def configuration(self) -> dict[str, Any]:
        return {"provider": self.name, "domain": self.domain, "configured": self.configured(), "endpoint": self.endpoint}


class AlphaVantageEquitiesAdapter(BaseAdapter):
    name = "alpha-vantage-global-quote"
    domain = "equities"
    endpoint = "https://www.alphavantage.co/query"
    key_env = "AURORA_ALPHA_VANTAGE_API_KEY"
    symbols_env = "AURORA_EQUITY_SYMBOLS"
    license_note = "Alpha Vantage terms, exchange entitlements and redistribution restrictions apply."
    completeness_note = "Coverage is limited to configured symbols. Quote freshness depends on entitlement and may be end-of-day or delayed."

    def api_key(self) -> str:
        return str(os.getenv(self.key_env) or "").strip()

    def symbols(self) -> list[str]:
        return [item.strip() for item in str(os.getenv(self.symbols_env) or "").split(",") if item.strip()][:50]

    def configured(self) -> bool:
        return bool(self.api_key() and self.symbols())

    def configuration(self) -> dict[str, Any]:
        return {
            "provider": self.name,
            "domain": self.domain,
            "configured": self.configured(),
            "key_env": self.key_env,
            "symbols_env": self.symbols_env,
            "symbol_count": len(self.symbols()),
            "credentials_never_returned": True,
        }

    def fetch(self, timeout: int = 30) -> list[MarketObservation]:
        if not self.configured():
            raise RuntimeError("Alpha Vantage equities provider is not configured")
        observed_at = _now()
        output: list[MarketObservation] = []
        for symbol in self.symbols():
            url = f"{self.endpoint}?{urllib.parse.urlencode({'function': 'GLOBAL_QUOTE', 'symbol': symbol, 'apikey': self.api_key()})}"
            try:
                payload = _json(url, timeout=timeout)
            except Exception as exc:
                raise RuntimeError(f"Alpha Vantage request failed for {symbol}") from exc
            quote = payload.get("Global Quote") if isinstance(payload, dict) else None
            if not isinstance(quote, dict):
                raise RuntimeError(f"Alpha Vantage returned no quote for {symbol}")
            price = _float(quote.get("05. price") or quote.get("price"))
            if price is None:
                raise RuntimeError(f"Alpha Vantage returned an invalid quote for {symbol}")
            event = _event_time(quote.get("07. latest trading day"), observed_at)
            output.append(
                MarketObservation(
                    self.domain,
                    self.name,
                    symbol,
                    f"{symbol}:{event}",
                    observed_at,
                    event,
                    price,
                    "provider quote currency",
                    "quote",
                    quote,
                    {
                        "provider": "Alpha Vantage",
                        "source": self.endpoint,
                        "credential_env": self.key_env,
                        "scope_env": self.symbols_env,
                        "credential_committed": False,
                        "request_url_persisted": False,
                        "license_note": self.license_note,
                    },
                )
            )
        return output


class EIAEnergyAdapter(BaseAdapter):
    name = "eia-api-v2"
    domain = "energy"
    endpoint = "https://api.eia.gov/v2"
    key_env = "AURORA_EIA_API_KEY"
    route_env = "AURORA_EIA_ROUTE"
    query_env = "AURORA_EIA_QUERY"
    value_env = "AURORA_EIA_VALUE_FIELD"
    license_note = "U.S. Energy Information Administration reuse policy and API terms apply."
    completeness_note = "The default scope is U.S. residential electricity prices in Colorado; operator routes may change scope and must be disclosed."

    def api_key(self) -> str:
        return str(os.getenv(self.key_env) or "").strip()

    def configured(self) -> bool:
        return bool(self.api_key())

    def configuration(self) -> dict[str, Any]:
        return {
            "provider": self.name,
            "domain": self.domain,
            "configured": self.configured(),
            "key_env": self.key_env,
            "route_env": self.route_env,
            "query_env": self.query_env,
            "value_env": self.value_env,
            "credentials_never_returned": True,
        }

    def fetch(self, timeout: int = 30) -> list[MarketObservation]:
        if not self.configured():
            raise RuntimeError("EIA energy provider is not configured")
        route = str(os.getenv(self.route_env) or "electricity/retail-sales/data/").strip("/")
        query = str(os.getenv(self.query_env) or "data[]=price&facets[sectorid][]=RES&facets[stateid][]=CO&frequency=monthly&sort[0][column]=period&sort[0][direction]=desc&length=12")
        value_field = str(os.getenv(self.value_env) or "price").strip()
        params = urllib.parse.parse_qsl(query, keep_blank_values=True)
        params.append(("api_key", self.api_key()))
        url = f"{self.endpoint}/{route}/?{urllib.parse.urlencode(params, doseq=True)}"
        try:
            payload = _json(url, timeout=timeout)
        except Exception as exc:
            raise RuntimeError("EIA API request failed") from exc
        rows = ((payload.get("response") or {}).get("data") or []) if isinstance(payload, dict) else []
        observed_at = _now()
        output: list[MarketObservation] = []
        for row in rows[:250]:
            if not isinstance(row, dict):
                continue
            value = _float(row.get(value_field))
            period = row.get("period")
            if value is None or period is None:
                continue
            instrument_parts = [route, str(row.get("stateid") or ""), str(row.get("sectorid") or "")]
            instrument = ":".join(part for part in instrument_parts if part)
            event = _event_time(period, observed_at)
            output.append(
                MarketObservation(
                    self.domain,
                    self.name,
                    instrument,
                    f"{instrument}:{event}:{value_field}",
                    observed_at,
                    event,
                    value,
                    str(row.get(f"{value_field}-units") or value_field),
                    "official-statistic",
                    row,
                    {
                        "provider": "U.S. Energy Information Administration",
                        "source": self.endpoint,
                        "credential_env": self.key_env,
                        "route_env": self.route_env,
                        "query_env": self.query_env,
                        "request_url_persisted": False,
                        "license_note": self.license_note,
                    },
                )
            )
        return output


class WorldBankCommoditiesAdapter(BaseAdapter):
    name = "world-bank-pink-sheet"
    domain = "commodities"
    endpoint = "https://thedocs.worldbank.org/en/doc/74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/CMO-Historical-Data-Monthly.xlsx"
    url_env = "AURORA_WORLD_BANK_COMMODITIES_URL"
    series_env = "AURORA_COMMODITY_SERIES"
    license_note = "World Bank commodity-price data terms and attribution requirements apply."
    completeness_note = "Monthly Pink Sheet prices are publication data, not real-time executable commodity quotes."

    def source_url(self) -> str:
        return str(os.getenv(self.url_env) or self.endpoint).strip()

    def selected(self) -> list[str]:
        raw = str(os.getenv(self.series_env) or "CRUDE_BRENT,GOLD,SILVER,COPPER,NGAS_US,COAL_AUS")
        return [item.strip().upper() for item in raw.split(",") if item.strip()][:50]

    def fetch(self, timeout: int = 45) -> list[MarketObservation]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for World Bank commodity ingestion") from exc
        try:
            content = _request(self.source_url(), timeout=timeout, accept="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise RuntimeError("World Bank Pink Sheet request or parse failed") from exc
        sheet = next((workbook[name] for name in workbook.sheetnames if "monthly" in name.lower()), workbook.active)
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next(
            (
                index
                for index, row in enumerate(rows[:30])
                if row and str(row[0] or "").strip().lower() in {"date", "month"}
            ),
            None,
        )
        if header_index is None:
            raise RuntimeError("World Bank Pink Sheet header was not found")
        headers = [str(value or "").strip() for value in rows[header_index]]
        units = [str(value or "").strip() for value in rows[header_index + 1]] if header_index + 1 < len(rows) else []
        data_rows = [row for row in rows[header_index + 1 :] if row and row[0] not in (None, "")]
        if not data_rows:
            raise RuntimeError("World Bank Pink Sheet returned no data rows")
        latest = data_rows[-1]
        observed_at = _now()
        event = _event_time(latest[0], observed_at)
        selected = set(self.selected())
        candidates: list[tuple[int, str]] = []
        for index, header in enumerate(headers[1:], start=1):
            normalized = header.upper().replace(" ", "_")
            if selected and not any(token == normalized or token in normalized for token in selected):
                continue
            if index < len(latest) and _float(latest[index]) is not None:
                candidates.append((index, header or f"series_{index}"))
        if not candidates:
            candidates = [
                (index, header or f"series_{index}")
                for index, header in enumerate(headers[1:], start=1)
                if index < len(latest) and _float(latest[index]) is not None
            ][:25]
        output: list[MarketObservation] = []
        for index, instrument in candidates[:50]:
            value = _float(latest[index])
            if value is None:
                continue
            unit = units[index] if index < len(units) and units[index] else "published unit"
            output.append(
                MarketObservation(
                    self.domain,
                    self.name,
                    instrument,
                    f"{instrument}:{event}",
                    observed_at,
                    event,
                    value,
                    unit,
                    "official-monthly-price",
                    {"series": instrument, "period": str(latest[0]), "value": value, "unit": unit},
                    {
                        "provider": "World Bank Prospects Group",
                        "source": self.source_url(),
                        "scope_env": self.series_env,
                        "license_note": self.license_note,
                    },
                )
            )
        return output


class ECBFXAdapter(BaseAdapter):
    name = "ecb-reference-rates"
    domain = "fx"
    endpoint = "https://data-api.ecb.europa.eu/service/data/EXR"
    currencies_env = "AURORA_ECB_CURRENCIES"
    license_note = "ECB Data Portal terms and attribution requirements apply."
    completeness_note = "ECB reference rates are informational averages against the euro, not executable global FX quotes."

    def currencies(self) -> list[str]:
        raw = str(os.getenv(self.currencies_env) or "USD,GBP,JPY,CHF,AUD,CAD,CNY")
        return [item.strip().upper() for item in raw.split(",") if item.strip()][:30]

    def fetch(self, timeout: int = 30) -> list[MarketObservation]:
        currencies = self.currencies()
        key = f"D.{'+'.join(currencies)}.EUR.SP00.A"
        url = f"{self.endpoint}/{key}?{urllib.parse.urlencode({'lastNObservations': 2, 'format': 'csvdata', 'detail': 'dataonly'})}"
        try:
            content = _request(url, timeout=timeout, accept="text/csv").decode("utf-8-sig", "replace")
        except Exception as exc:
            raise RuntimeError("ECB exchange-rate request failed") from exc
        observed_at = _now()
        output: list[MarketObservation] = []
        for row in csv.DictReader(io.StringIO(content)):
            currency = str(row.get("CURRENCY") or row.get("currency") or "").strip().upper()
            value = _float(row.get("OBS_VALUE") or row.get("obs_value"))
            period = row.get("TIME_PERIOD") or row.get("time_period")
            if not currency or value is None or not period:
                continue
            event = _event_time(period, observed_at)
            instrument = f"EUR/{currency}"
            output.append(
                MarketObservation(
                    self.domain,
                    self.name,
                    instrument,
                    f"{instrument}:{event}",
                    observed_at,
                    event,
                    value,
                    f"{currency} per EUR",
                    "reference-rate",
                    {"currency": currency, "period": period, "value": value},
                    {"provider": "European Central Bank", "source": self.endpoint, "scope_env": self.currencies_env, "license_note": self.license_note},
                )
            )
        return output


class CoinbaseCryptoAdapter(BaseAdapter):
    name = "coinbase-exchange-ticker"
    domain = "crypto"
    endpoint = "https://api.exchange.coinbase.com/products"
    products_env = "AURORA_COINBASE_PRODUCTS"
    license_note = "Coinbase Exchange API terms and market-data restrictions apply."
    completeness_note = "Coverage is limited to configured Coinbase products and one venue; it is not a consolidated crypto market."

    def products(self) -> list[str]:
        raw = str(os.getenv(self.products_env) or "BTC-USD,ETH-USD")
        return [item.strip().upper() for item in raw.split(",") if item.strip()][:50]

    def fetch(self, timeout: int = 30) -> list[MarketObservation]:
        observed_at = _now()
        output: list[MarketObservation] = []
        for product in self.products():
            url = f"{self.endpoint}/{urllib.parse.quote(product, safe='-')}/ticker"
            try:
                payload = _json(url, timeout=timeout)
            except Exception as exc:
                raise RuntimeError(f"Coinbase ticker request failed for {product}") from exc
            price = _float(payload.get("price")) if isinstance(payload, dict) else None
            event = _event_time(payload.get("time") if isinstance(payload, dict) else None, observed_at)
            if price is None:
                raise RuntimeError(f"Coinbase returned an invalid ticker for {product}")
            quote_currency = product.split("-")[-1] if "-" in product else "quote currency"
            output.append(
                MarketObservation(
                    self.domain,
                    self.name,
                    product,
                    f"{product}:{event}",
                    observed_at,
                    event,
                    price,
                    quote_currency,
                    "venue-ticker",
                    payload,
                    {"provider": "Coinbase Exchange", "source": url, "scope_env": self.products_env, "license_note": self.license_note},
                )
            )
        return output


class WorldBankIndicatorsAdapter(BaseAdapter):
    name = "world-bank-indicators-v2"
    domain = "economic_indicators"
    endpoint = "https://api.worldbank.org/v2"
    countries_env = "AURORA_WORLD_BANK_COUNTRIES"
    indicators_env = "AURORA_WORLD_BANK_INDICATORS"
    license_note = "World Bank Open Data terms and source-specific attribution requirements apply."
    completeness_note = "World Bank indicators are official publication series with varying lags and revisions, not real-time macroeconomic releases."

    def countries(self) -> list[str]:
        raw = str(os.getenv(self.countries_env) or "USA,CHN,IND,DEU,JPN,GBR")
        return [item.strip().upper() for item in raw.replace(";", ",").split(",") if item.strip()][:60]

    def indicators(self) -> list[str]:
        raw = str(os.getenv(self.indicators_env) or "NY.GDP.MKTP.CD,FP.CPI.TOTL.ZG,SL.UEM.TOTL.ZS")
        return [item.strip().upper() for item in raw.split(",") if item.strip()][:30]

    def fetch(self, timeout: int = 30) -> list[MarketObservation]:
        observed_at = _now()
        country_path = ";".join(self.countries())
        output: list[MarketObservation] = []
        for indicator in self.indicators():
            url = f"{self.endpoint}/country/{country_path}/indicator/{urllib.parse.quote(indicator, safe='.') }?{urllib.parse.urlencode({'format': 'json', 'per_page': 1000, 'mrnev': 1})}"
            try:
                payload = _json(url, timeout=timeout)
            except Exception as exc:
                raise RuntimeError(f"World Bank indicator request failed for {indicator}") from exc
            rows = payload[1] if isinstance(payload, list) and len(payload) > 1 and isinstance(payload[1], list) else []
            for row in rows:
                value = _float(row.get("value")) if isinstance(row, dict) else None
                if value is None:
                    continue
                country = str((row.get("country") or {}).get("id") or row.get("countryiso3code") or "").upper()
                code = str((row.get("indicator") or {}).get("id") or indicator)
                period = str(row.get("date") or "")
                if not country or not period:
                    continue
                event = _event_time(period, observed_at)
                instrument = f"{country}:{code}"
                output.append(
                    MarketObservation(
                        self.domain,
                        self.name,
                        instrument,
                        f"{instrument}:{event}",
                        observed_at,
                        event,
                        value,
                        "indicator-defined unit",
                        "official-statistic",
                        row,
                        {"provider": "World Bank", "source": url, "countries_env": self.countries_env, "indicators_env": self.indicators_env, "license_note": self.license_note},
                    )
                )
        return output


class KalshiPredictionMarketsAdapter(BaseAdapter):
    name = "kalshi-public-markets"
    domain = "prediction_markets"
    endpoint = "https://external-api.kalshi.com/trade-api/v2/markets"
    limit_env = "AURORA_KALSHI_MARKET_LIMIT"
    license_note = "Kalshi Developer Agreement, market-data terms and applicable legal restrictions apply."
    completeness_note = "Coverage is limited to open Kalshi markets on one regulated venue and is not a universal probability forecast."

    def fetch(self, timeout: int = 30) -> list[MarketObservation]:
        limit = max(1, min(int(os.getenv(self.limit_env) or "250"), 1000))
        url = f"{self.endpoint}?{urllib.parse.urlencode({'status': 'open', 'limit': limit})}"
        try:
            payload = _json(url, timeout=timeout)
        except Exception as exc:
            raise RuntimeError("Kalshi public market request failed") from exc
        observed_at = _now()
        output: list[MarketObservation] = []
        for row in (payload.get("markets") or []) if isinstance(payload, dict) else []:
            ticker = str(row.get("ticker") or "").strip()
            if not ticker:
                continue
            value = _float(row.get("last_price_dollars"))
            bid = _float(row.get("yes_bid_dollars"))
            ask = _float(row.get("yes_ask_dollars"))
            if value is None and bid is not None and ask is not None:
                value = (bid + ask) / 2.0
            if value is None:
                continue
            event = _event_time(row.get("updated_time") or row.get("open_time") or row.get("created_time"), observed_at)
            output.append(
                MarketObservation(
                    self.domain,
                    self.name,
                    ticker,
                    f"{ticker}:{event}",
                    observed_at,
                    event,
                    value,
                    "yes probability in dollars",
                    str(row.get("status") or "open"),
                    row,
                    {"provider": "Kalshi", "source": self.endpoint, "market_scope": "open", "license_note": self.license_note},
                )
            )
        return output


class MarketCoordinator:
    def __init__(self, store: MarketStore) -> None:
        self.store = store
        self.adapters: list[BaseAdapter] = [
            AlphaVantageEquitiesAdapter(),
            EIAEnergyAdapter(),
            WorldBankCommoditiesAdapter(),
            ECBFXAdapter(),
            CoinbaseCryptoAdapter(),
            WorldBankIndicatorsAdapter(),
            KalshiPredictionMarketsAdapter(),
        ]
        self._ensure_registered()

    def _ensure_registered(self) -> None:
        existing = {row["provider"] for row in self.store.providers()}
        for adapter in self.adapters:
            if adapter.name not in existing:
                self.store.upsert_provider(
                    {
                        "provider": adapter.name,
                        "domain": adapter.domain,
                        "state": "NOT_CONFIGURED",
                        "last_attempt_at": "",
                        "last_success_at": "",
                        "consecutive_failures": 0,
                        "event_age_seconds": 0,
                        "last_error": "",
                        "completeness_note": adapter.completeness_note,
                        "license_note": adapter.license_note,
                    }
                )

    def _observe(self, adapter: BaseAdapter, *, successful: bool, configured: bool, event_age_seconds: int, error: str) -> None:
        existing = next(row for row in self.store.providers() if row["provider"] == adapter.name)
        now = _now()
        failures = 0 if successful else int(existing.get("consecutive_failures") or 0) + (1 if configured else 0)
        if not configured:
            state = "NOT_CONFIGURED"
        elif successful:
            state = "ONLINE"
        else:
            state = "OFFLINE" if failures >= 3 else "DEGRADED"
        self.store.upsert_provider(
            {
                **existing,
                "state": state,
                "last_attempt_at": now if configured else existing.get("last_attempt_at", ""),
                "last_success_at": now if successful else existing.get("last_success_at", ""),
                "consecutive_failures": failures,
                "event_age_seconds": max(0, int(event_age_seconds)),
                "last_error": "" if successful or not configured else error,
            }
        )

    def run(self, provider: str, *, timeout: int = 30) -> ProviderRun:
        adapter = next((row for row in self.adapters if row.name == provider), None)
        if adapter is None:
            raise KeyError("market provider not found")
        started_at = _now()
        started_clock = time.monotonic()
        configured = adapter.configured()
        observations: list[MarketObservation] = []
        error = ""
        if configured:
            try:
                observations = adapter.fetch(timeout=timeout)
                if not observations:
                    error = "provider returned no valid numeric observations"
            except Exception as exc:
                error = str(exc)
        else:
            error = "provider is not configured"
        successful = configured and bool(observations) and not error
        event_age = 0
        if observations:
            times = [_parse_time(row.event_time) for row in observations]
            valid = [value for value in times if value is not None]
            event_age = max(0, int((datetime.now(timezone.utc) - max(valid)).total_seconds())) if valid else 0
        if successful:
            for observation in observations:
                self.store.record(observation)
        completed_at = _now()
        run = ProviderRun(
            adapter.name,
            adapter.domain,
            configured,
            successful,
            len(observations),
            error,
            started_at,
            completed_at,
            max(0, int((time.monotonic() - started_clock) * 1000)),
        )
        self._observe(adapter, successful=successful, configured=configured, event_age_seconds=event_age, error=error)
        self.store.record_run(run, {"event_age_seconds": event_age})
        return run

    def run_all(self, *, timeout: int = 30) -> list[ProviderRun]:
        return [self.run(adapter.name, timeout=timeout) for adapter in self.adapters]

    def configuration(self) -> dict[str, Any]:
        return {
            "providers": [adapter.configuration() for adapter in self.adapters],
            "credentials_never_returned": True,
            "registration_is_not_live_evidence": True,
            "retrieval_is_not_investment_advice": True,
        }
